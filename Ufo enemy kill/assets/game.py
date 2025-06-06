import math
import random
import os
import pygame
from pygame import mixer
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import atexit

# =================================================================
#                         INITIALIZATION
# =================================================================

pygame.mixer.pre_init(44100, -16, 2, 512)
pygame.init()
FPS = 60
clock = pygame.time.Clock()

# --- Screen Setup ---
SCREEN_WIDTH, SCREEN_HEIGHT = 1100, 700
screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))
pygame.display.set_caption("Space Invaders: Final Frontier")

# --- Asset Paths ---
ASSET_PATH = r"D:\Ufo enemy kill\assets"

# --- Load Assets ---
try:
    original_bg = pygame.image.load(os.path.join(ASSET_PATH, 'background.png'))
    background_img = pygame.transform.scale(original_bg, (SCREEN_WIDTH, SCREEN_HEIGHT))
    
    icon_img = pygame.image.load(os.path.join(ASSET_PATH, 'ufo.png'))
    pygame.display.set_icon(icon_img)
    player_img = pygame.image.load(os.path.join(ASSET_PATH, 'player.png'))
    enemy_img = pygame.image.load(os.path.join(ASSET_PATH, 'enemy.png'))
    bullet_img = pygame.image.load(os.path.join(ASSET_PATH, 'bullet.png'))

    mixer.music.load(os.path.join(ASSET_PATH, 'background_music.wav'))
    mixer.music.set_volume(0.3)
    bullet_sound = mixer.Sound(os.path.join(ASSET_PATH, 'laser.wav'))
    explosion_sound = mixer.Sound(os.path.join(ASSET_PATH, 'explosion.wav'))

except pygame.error as e:
    print(f"FATAL ERROR: Asset loading failed. Ensure path '{ASSET_PATH}' is correct. Details: {e}")
    exit()

# --- Fonts ---
try:
    emoji_font = pygame.font.Font("C:/Windows/Fonts/seguiemj.ttf", 30)
except FileNotFoundError:
    emoji_font = pygame.font.Font(None, 40)
score_font = pygame.font.Font(None, 32)
game_over_font = pygame.font.Font(None, 64)
info_font = pygame.font.Font(None, 24)
input_font = pygame.font.Font(None, 40)
title_font = pygame.font.Font(None, 80)
danger_font = pygame.font.Font(None, 40)

# --- Colors & UI Elements ---
WHITE, BLACK = (255, 255, 255), (0, 0, 0)
GREEN, RED, BLUE = (0, 200, 0), (200, 0, 0), (50, 50, 200)
GOLD = (255, 215, 0)
quit_button_rect = pygame.Rect(SCREEN_WIDTH - 110, SCREEN_HEIGHT - 50, 100, 40)
pause_button_rect = pygame.Rect(SCREEN_WIDTH - 60, 5, 50, 35)

# =================================================================
#                      GAME STATE VARIABLES
# =================================================================
player_x_change, player_y_change = 0, 0
current_sensitivity_level = 3
sensitivity_map = {1: 3, 2: 4, 3: 6, 4: 8, 5: 11}
player_rect = player_img.get_rect()
bullet_rect = bullet_img.get_rect()

user_name, user_gender = "", ""
score_value, lives = 0, 3
bullet_state, is_paused, sound_enabled = "ready", False, True
game_session_active = False # To control atexit saving
enemies, danger_level, enemy_base_speed = [], 1, 3.5

# =================================================================
#                     DATA & HELPER FUNCTIONS
# =================================================================

def save_to_excel(outcome):
    if not user_name: return # Do not save if no user is set
    filename = "Ufo Game.xlsx"
    headers = ["Player Name", "Gender", "Final Score", "Danger Level", "Game Outcome", "Timestamp"]
    try:
        workbook = load_workbook(filename) if os.path.exists(filename) else openpyxl.Workbook()
        sheet = workbook.active
        if sheet.max_row == 0 or (sheet.max_row == 1 and sheet.cell(1,1).value is None):
            sheet.title = "Game Report"; sheet.append(headers)
        data_row = [user_name, user_gender, score_value, danger_level, outcome, datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        sheet.append(data_row)
        for col in sheet.columns:
            max_length = 0; column_letter = col[0].column_letter
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            sheet.column_dimensions[column_letter].width = max_length + 2
        workbook.save(filename)
        print(f"Game data saved to '{filename}'")
    except Exception as e:
        print(f"Error saving to Excel: {e}")

def save_on_exit():
    if game_session_active: save_to_excel("Terminated")
atexit.register(save_on_exit)


def reset_game_state():
    global player_x_change, player_y_change, score_value, lives, bullet_state, enemies, danger_level
    player_rect.centerx, player_rect.bottom = SCREEN_WIDTH // 2, SCREEN_HEIGHT - 60
    player_x_change, player_y_change = 0, 0
    score_value, lives, danger_level = 0, 3, 1
    bullet_state = "ready"
    enemies = [create_enemy() for _ in range(8)]

def create_enemy():
    rect = enemy_img.get_rect(x=random.randint(0, SCREEN_WIDTH - 64), y=random.randint(50, 200))
    return {"rect": rect, "x_speed": random.choice([-1, 1]), "behavior": "normal"}

def play_sound(sound):
    if sound_enabled: sound.play()

def get_high_score():
    filename = "Ufo Game.xlsx"
    high_score, high_scorer = 0, "N/A"
    if not os.path.exists(filename): return high_score, high_scorer
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] is not None and isinstance(row[2], (int, float)) and row[2] > high_score:
                high_score, high_scorer = row[2], row[0]
    except Exception: pass
    return high_score, high_scorer

# =================================================================
#                         UI & DRAWING
# =================================================================

def draw_text(text, font, color, surface, x, y, center=True):
    textobj = font.render(text, 1, color)
    textrect = textobj.get_rect(center=(x, y)) if center else textobj.get_rect(topleft=(x, y))
    surface.blit(textobj, textrect)

def draw_button(rect, text, base_color, highlight_color):
    color = highlight_color if rect.collidepoint(pygame.mouse.get_pos()) else base_color
    pygame.draw.rect(screen, color, rect, border_radius=10)
    draw_text(text, score_font, WHITE, screen, rect.centerx, rect.centery)

def show_game_ui():
    draw_text(f"Score: {score_value}", score_font, WHITE, screen, 80, 25)
    draw_text(f"DANGER LEVEL: {danger_level}", danger_font, GOLD, screen, SCREEN_WIDTH / 2, 25)
    lives_text = " ".join(['❤️'] * lives); lives_display = emoji_font.render(lives_text, True, RED)
    screen.blit(lives_display, (SCREEN_WIDTH - lives_display.get_width() - 80, 5))
    draw_button(quit_button_rect, 'Quit', RED, (255, 50, 50))
    pygame.draw.rect(screen, BLUE, pause_button_rect, border_radius=8)
    draw_text("||", score_font, WHITE, screen, pause_button_rect.centerx, pause_button_rect.centery)

# =================================================================
#                            SCREENS
# =================================================================

def home_screen():
    global user_name, user_gender, current_sensitivity_level, sound_enabled
    high_score, high_scorer = get_high_score()
    name_box = pygame.Rect(SCREEN_WIDTH/2 - 200, 250, 400, 50)
    male_box = pygame.Rect(SCREEN_WIDTH/2 - 200, 330, 190, 50)
    female_box = pygame.Rect(SCREEN_WIDTH/2 + 10, 330, 190, 50)
    start_box = pygame.Rect(SCREEN_WIDTH/2 - 150, 410, 300, 60)
    sensitivity_slider = pygame.Rect(SCREEN_WIDTH/2 - 150, 520, 300, 30)
    sound_toggle_box = pygame.Rect(SCREEN_WIDTH/2 - 150, 560, 300, 40)
    quit_box = pygame.Rect(SCREEN_WIDTH/2 - 150, 610, 300, 50)
    name_active = False; user_name, user_gender = "", ""

    while True:
        screen.blit(background_img, (0, 0))
        draw_text("Space Invaders: Final Frontier", title_font, WHITE, screen, SCREEN_WIDTH / 2, 70)
        draw_text(f"High Score: {high_score} by {high_scorer}", score_font, GOLD, screen, SCREEN_WIDTH/2, 150)
        draw_text("Controls: WASD/Arrows | P-Pause | U-Unpause | SPACE-Shoot", info_font, WHITE, screen, SCREEN_WIDTH/2, 190)

        for event in pygame.event.get():
            if event.type == pygame.QUIT: pygame.quit(); exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                name_active = name_box.collidepoint(event.pos)
                if male_box.collidepoint(event.pos): user_gender = "Male"
                elif female_box.collidepoint(event.pos): user_gender = "Female"
                elif start_box.collidepoint(event.pos) and user_name and user_gender: return
                elif quit_box.collidepoint(event.pos): pygame.quit(); exit()
                elif sound_toggle_box.collidepoint(event.pos):
                    sound_enabled = not sound_enabled
                    if not sound_enabled: mixer.music.stop()
                if sensitivity_slider.collidepoint(event.pos):
                    current_sensitivity_level = min(5, max(1, int(((event.pos[0] - sensitivity_slider.x) / sensitivity_slider.width) * 5) + 1))
            if event.type == pygame.KEYDOWN and name_active:
                user_name = user_name[:-1] if event.key == pygame.K_BACKSPACE else user_name + event.unicode

        draw_button(start_box, 'START GAME', GREEN if user_name and user_gender else (50,50,50), (100,255,100))
        draw_button(male_box, 'Male', BLUE if user_gender == "Male" else (50,50,50), (100,100,255))
        draw_button(female_box, 'Female', (200,0,100) if user_gender == "Female" else (50,50,50), (255,100,180))
        draw_button(quit_box, 'QUIT', RED, (255,50,50))
        draw_button(sound_toggle_box, f"Sound: {'ON' if sound_enabled else 'OFF'}", GREEN if sound_enabled else RED, (100,255,100) if sound_enabled else (255,100,100))

        pygame.draw.rect(screen, pygame.Color('lightskyblue3') if name_active else WHITE, name_box, 2, border_radius=10)
        draw_text(user_name, input_font, WHITE, screen, name_box.x + 10, name_box.centery-15, center=False)
        draw_text("Movement Sensitivity", info_font, WHITE, screen, SCREEN_WIDTH/2, 500)
        pygame.draw.rect(screen, WHITE, sensitivity_slider, 2, border_radius=10); slider_fill_width = (current_sensitivity_level / 5) * sensitivity_slider.width
        pygame.draw.rect(screen, GREEN, (sensitivity_slider.x, sensitivity_slider.y, slider_fill_width, sensitivity_slider.height), border_radius=10)

        pygame.display.update(); clock.tick(FPS)

def pause_screen():
    global is_paused
    if sound_enabled: mixer.music.pause()
    while is_paused:
        for event in pygame.event.get():
            if event.type == pygame.QUIT: pygame.quit(); exit()
            if event.type == pygame.KEYDOWN and event.key == pygame.K_u: is_paused = False
            if event.type == pygame.MOUSEBUTTONDOWN and pause_button_rect.collidepoint(event.pos): is_paused = False

        draw_text("PAUSED", game_over_font, GOLD, screen, SCREEN_WIDTH/2, SCREEN_HEIGHT/2 - 40)
        draw_text("Press 'U' or click the pause icon to unpause", info_font, WHITE, screen, SCREEN_WIDTH/2, SCREEN_HEIGHT/2 + 20)
        pygame.display.update()
        clock.tick(15)
    if sound_enabled: mixer.music.unpause()

def game_loop():
    global player_x_change, player_y_change, bullet_state, score_value, lives, danger_level, is_paused, game_session_active
    reset_game_state(); game_session_active = True
    if sound_enabled: mixer.music.play(-1)

    difficulty_timer = pygame.time.get_ticks()
    dive_timer = pygame.time.get_ticks()

    running = True
    while running:
        if is_paused: pause_screen()

        screen.blit(background_img, (0, 0))

        if pygame.time.get_ticks() - difficulty_timer > 30000:
            danger_level += 1; difficulty_timer = pygame.time.get_ticks()
        if pygame.time.get_ticks() - dive_timer > 5000:
            if enemies: # Make sure enemy list is not empty
                random_enemy = random.choice(enemies)
                if random_enemy['behavior'] == 'normal': random_enemy['behavior'] = 'diving'
            dive_timer = pygame.time.get_ticks()

        player_speed = sensitivity_map[current_sensitivity_level]
        for event in pygame.event.get():
            if event.type == pygame.QUIT: pygame.quit(); exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if quit_button_rect.collidepoint(event.pos): save_to_excel("Quit"); running = False
                if pause_button_rect.collidepoint(event.pos): is_paused = True
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_p: is_paused = True
                if event.key in (pygame.K_LEFT, pygame.K_a): player_x_change = -player_speed
                if event.key in (pygame.K_RIGHT, pygame.K_d): player_x_change = player_speed
                if event.key in (pygame.K_UP, pygame.K_w): player_y_change = -player_speed
                if event.key in (pygame.K_DOWN, pygame.K_s): player_y_change = player_speed
                if event.key == pygame.K_SPACE and bullet_state == "ready":
                    play_sound(bullet_sound); bullet_state = "fire"; bullet_rect.center = player_rect.center
            if event.type == pygame.KEYUP:
                if event.key in (pygame.K_LEFT, pygame.K_a, pygame.K_RIGHT, pygame.K_d): player_x_change = 0
                if event.key in (pygame.K_UP, pygame.K_w, pygame.K_DOWN, pygame.K_s): player_y_change = 0

        player_rect.x += player_x_change; player_rect.y += player_y_change
        player_rect.clamp_ip(screen.get_rect())

        for enemy in list(enemies):
            current_enemy_speed = enemy_base_speed + (danger_level * 0.5)
            if enemy['behavior'] == 'normal':
                enemy['rect'].x += enemy['x_speed'] * current_enemy_speed
                if enemy['rect'].left <= 0 or enemy['rect'].right >= SCREEN_WIDTH: enemy['x_speed'] *= -1; enemy['rect'].y += 30
            elif enemy['behavior'] == 'diving':
                enemy['rect'].y += current_enemy_speed * 1.5
                if enemy['rect'].top > SCREEN_HEIGHT: enemies.remove(enemy); enemies.append(create_enemy())
            if bullet_state == "fire" and enemy['rect'].colliderect(bullet_rect):
                play_sound(explosion_sound); bullet_state = "ready"; score_value += 10
                enemies.remove(enemy); enemies.append(create_enemy())
            if player_rect.colliderect(enemy['rect']):
                lives -= 1; enemies.remove(enemy); enemies.append(create_enemy())
                if lives <= 0: save_to_excel("Game Over"); running = False

        if bullet_state == "fire":
            bullet_rect.y -= 15
            if bullet_rect.bottom < 0: bullet_state = "ready"

        screen.blit(player_img, player_rect)
        for enemy in enemies: screen.blit(enemy_img, enemy['rect'])
        if bullet_state == "fire": screen.blit(bullet_img, bullet_rect)
        show_game_ui()
        pygame.display.update()
        clock.tick(FPS)

    mixer.music.stop(); game_session_active = False

    if lives <= 0:
        waiting_for_input = True
        while waiting_for_input:
            screen.blit(background_img, (0, 0))
            draw_text("GAME OVER", game_over_font, WHITE, screen, SCREEN_WIDTH/2, SCREEN_HEIGHT/2 - 50)
            draw_text("Press ENTER to return to Home", score_font, WHITE, screen, SCREEN_WIDTH/2, SCREEN_HEIGHT/2 + 20)
            for event in pygame.event.get():
                if event.type == pygame.QUIT: pygame.quit(); exit()
                if event.type == pygame.KEYDOWN and event.key == pygame.K_RETURN: waiting_for_input = False
            pygame.display.update()
            clock.tick(15)

if __name__ == "__main__":
    while True:
        home_screen()
        game_loop()