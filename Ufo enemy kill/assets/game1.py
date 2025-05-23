import atexit
import math
import random
import os
import pygame
from pygame import mixer
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

# =================================================================
#                         INITIALIZATION
# =================================================================

pygame.init()

# --- Screen Setup ---
SCREEN_WIDTH, SCREEN_HEIGHT = 1100, 700
screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))
pygame.display.set_caption("Space Invaders Deluxe")

# --- Asset Paths ---
ASSET_PATH = r"D:\Ufo enemy kill\assets"

# --- Load Assets ---
try:
    original_bg = pygame.image.load(os.path.join(ASSET_PATH, 'background.png'))
    background_img = pygame.transform.scale(original_bg, (SCREEN_WIDTH, SCREEN_HEIGHT))

    icon_img = pygame.image.load(os.path.join(ASSET_PATH, 'ufo.png'))
    pygame.display.set_icon(icon_img)
    player_img = pygame.image.load(os.path.join(ASSET_PATH, 'player.png'))
    enemy_img = pygame.image.load(os.path.join(ASSET_PATH, 'enemy.png'))
    bullet_img = pygame.image.load(os.path.join(ASSET_PATH, 'bullet.png'))
    
    # Sounds
    mixer.music.load(os.path.join(ASSET_PATH, 'background_music.wav'))
    mixer.music.set_volume(0.3)
    bullet_sound = mixer.Sound(os.path.join(ASSET_PATH, 'laser.wav'))
    explosion_sound = mixer.Sound(os.path.join(ASSET_PATH, 'explosion.wav'))

except pygame.error as e:
    print("---------------------------------------------------------")
    print(f"FATAL ERROR: Asset loading failed.")
    print(f"Please ensure the path '{ASSET_PATH}' is correct and contains all required files.")
    print(f"Details: {e}")
    print("---------------------------------------------------------")
    exit()

# --- Fonts ---
try:
    emoji_font = pygame.font.Font("C:/Windows/Fonts/seguiemj.ttf", 30)
except FileNotFoundError:
    emoji_font = pygame.font.Font(None, 40)
score_font = pygame.font.Font(None, 32)
game_over_font = pygame.font.Font(None, 64)
info_font = pygame.font.Font(None, 24)
input_font = pygame.font.Font(None, 40)
title_font = pygame.font.Font(None, 80)

# --- Colors & Buttons ---
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
GREEN = (0, 200, 0)
RED = (200, 0, 0)
BLUE = (50, 50, 200)
quit_button_rect = pygame.Rect(SCREEN_WIDTH - 110, SCREEN_HEIGHT - 50, 100, 40)

# =================================================================
#                      GAME STATE VARIABLES
# =================================================================

# Player movement
player_x_change = 0
player_y_change = 0
current_sensitivity = 3  # Default sensitivity (1-5 scale)

# Player
player_x = (SCREEN_WIDTH - 64) // 2
player_y = SCREEN_HEIGHT - 100  # Start at bottom

# Enemies
enemies = []
num_of_enemies = 6
enemy_speed_multiplier = 1.0

# Bullet
bullet_x, bullet_y = 0, 0
bullet_state = "ready"
bullet_y_change = 10

# Game Stats
score_value = 0
lives = 3
user_name = ""
user_gender = ""

# =================================================================
#                     DATA & HELPER FUNCTIONS
# =================================================================

def save_to_excel(outcome):
    """Saves game data to a single Excel file, creating it if it doesn't exist"""
    filename = "Ufo Game.xlsx"
    
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
        else:
            workbook = openpyxl.Workbook()
        
        if "Game Report" in workbook.sheetnames:
            sheet = workbook["Game Report"]
        else:
            sheet = workbook.active
            sheet.title = "Game Report"
            headers = ["Player Name", "Gender", "Final Score", "Game Outcome", "Timestamp"]
            sheet.append(headers)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data_row = [user_name, user_gender, score_value, outcome, timestamp]
        sheet.append(data_row)
        
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(filename)
        print(f"Game data successfully saved to '{filename}'")

    except Exception as e:
        print(f"Error: Could not save data to Excel file. {e}")

def save_on_exit():
    """Function to be called when the game is exiting"""
    if 'user_name' in globals() and user_name:
        save_to_excel("Terminated")

def reset_game_state():
    """Resets all variables for a new game session."""
    global player_x, player_y, player_x_change, player_y_change, score_value, lives
    global enemies, bullet_state, enemy_speed_multiplier

    player_x = (SCREEN_WIDTH - 64) // 2
    player_y = SCREEN_HEIGHT - 100  # Start at bottom
    player_x_change = 0
    player_y_change = 0
    score_value = 0
    lives = 3
    bullet_state = "ready"
    enemy_speed_multiplier = 1.0

    enemies.clear()
    for _ in range(num_of_enemies):
        enemies.append({
            "x": random.randint(0, SCREEN_WIDTH - 64),
            "y": random.randint(50, 150),
            "x_change": 4 * enemy_speed_multiplier
        })

    bullet_x = 0
    bullet_y = 0

def create_enemy():
    """Creates a new enemy at a random position."""
    return {
        "x": random.randint(0, SCREEN_WIDTH - 64),
        "y": random.randint(50, 150),
        "x_change": 4 * enemy_speed_multiplier
    }

def draw_text(text, font, color, surface, x, y):
    """Helper function to draw text on a surface."""
    textobj = font.render(text, 1, color)
    textrect = textobj.get_rect(center=(x, y))
    surface.blit(textobj, textrect)

def draw_button(rect, text, color):
    """Draws a clickable button."""
    pygame.draw.rect(screen, color, rect)
    draw_text(text, info_font, WHITE, screen, rect.centerx, rect.centery)

def show_game_ui():
    """Renders all the in-game UI elements."""
    # Score
    score_display = score_font.render(f"Score: {score_value}", True, WHITE)
    screen.blit(score_display, (10, 10))
    
    # Lives
    lives_text = " ".join(['❤️'] * lives)
    lives_display = emoji_font.render(lives_text, True, RED)
    screen.blit(lives_display, (SCREEN_WIDTH - lives_display.get_width() - 10, 5))

    # Quit Button
    draw_button(quit_button_rect, 'Quit Game', RED)

def game_over_screen():
    """Displays the final game over message."""
    mixer.music.stop()
    draw_text("GAME OVER", game_over_font, WHITE, screen, SCREEN_WIDTH / 2, SCREEN_HEIGHT / 2 - 50)
    draw_text(f"Final Score: {score_value}", score_font, WHITE, screen, SCREEN_WIDTH / 2, SCREEN_HEIGHT / 2 + 20)
    draw_text("Press ENTER to return to Home", info_font, WHITE, screen, SCREEN_WIDTH / 2, SCREEN_HEIGHT / 2 + 70)

def get_high_score():
    """Reads the Excel file and returns the highest score and scorer name"""
    filename = "Ufo Game.xlsx"
    high_score = 0
    high_scorer = "Empty"
    
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
            if "Game Report" in workbook.sheetnames:
                sheet = workbook["Game Report"]
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 3 and isinstance(row[2], (int, float)):
                        if row[2] > high_score:
                            high_score = row[2]
                            high_scorer = row[0] if row[0] else "Unknown"
    except Exception as e:
        print(f"Error reading high score: {e}")
    
    return high_score, high_scorer

# =================================================================
#                           HOME SCREEN
# =================================================================

def home_screen():
    """Manages the user input and selection screen with high score display and sensitivity controls"""
    global user_name, user_gender, current_sensitivity

    high_score, high_scorer = get_high_score()
    
    # UI Layout
    title_y = 80
    high_score_y = 160
    input_start_y = 280
    sensitivity_y = input_start_y + 400

    # UI Elements
    name_box = pygame.Rect(SCREEN_WIDTH/2 - 200, input_start_y, 400, 60)
    male_box = pygame.Rect(SCREEN_WIDTH/2 - 200, input_start_y + 120, 180, 60)
    female_box = pygame.Rect(SCREEN_WIDTH/2 + 20, input_start_y + 120, 180, 60)
    start_box = pygame.Rect(SCREEN_WIDTH/2 - 150, input_start_y + 220, 300, 70)
    quit_box = pygame.Rect(SCREEN_WIDTH/2 - 150, input_start_y + 320, 300, 60)
    sensitivity_box = pygame.Rect(SCREEN_WIDTH/2 - 200, sensitivity_y, 400, 30)

    # Colors
    active_color = pygame.Color('lightskyblue3')
    passive_color = pygame.Color('gray15')
    name_color = passive_color
    sensitivity_colors = [
        (100, 100, 100),  # Level 1
        (150, 150, 150),  # Level 2
        (200, 200, 200),  # Level 3
        (0, 200, 0),      # Level 4
        (0, 255, 0)       # Level 5
    ]

    name_active = False
    user_name = ""
    user_gender = ""

    while True:
        screen.blit(background_img, (0, 0))
        
        # Title
        draw_text("SPACE INVADERS DELUXE", title_font, WHITE, screen, SCREEN_WIDTH / 2, title_y)
        
        # High Score Display
        draw_text("THE HIGH SCORE IS:", score_font, WHITE, screen, SCREEN_WIDTH / 2 - 200, high_score_y)
        draw_text(str(high_score), score_font, (255, 215, 0), screen, SCREEN_WIDTH / 2 - 200, high_score_y + 40)
        draw_text("HIGH SCORER NAME:", score_font, WHITE, screen, SCREEN_WIDTH / 2 + 200, high_score_y)
        draw_text(high_scorer, score_font, (255, 215, 0), screen, SCREEN_WIDTH / 2 + 200, high_score_y + 40)

        # Controls Info
        controls_text = "CONTROLS: WASD or Arrow Keys to Move | SPACE to Shoot"
        draw_text(controls_text, info_font, WHITE, screen, SCREEN_WIDTH / 2, high_score_y + 100)

        # Name Input
        draw_text("ENTER YOUR NAME:", info_font, WHITE, screen, SCREEN_WIDTH / 2, input_start_y - 40)

        # Gender Selection
        draw_text("SELECT GENDER:", info_font, WHITE, screen, SCREEN_WIDTH / 2, input_start_y + 80)

        # Sensitivity Control
        draw_text("MOVEMENT SENSITIVITY:", info_font, WHITE, screen, SCREEN_WIDTH / 2, sensitivity_y - 30)

        # Event Handling
        mouse_pos = pygame.mouse.get_pos()
        
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                exit()

            if event.type == pygame.MOUSEBUTTONDOWN:
                if name_box.collidepoint(event.pos):
                    name_active = True
                else:
                    name_active = False
                
                if male_box.collidepoint(event.pos):
                    user_gender = "Male"
                elif female_box.collidepoint(event.pos):
                    user_gender = "Female"
                elif start_box.collidepoint(event.pos) and user_name and user_gender:
                    return
                elif quit_box.collidepoint(event.pos):
                    pygame.quit()
                    exit()
                    
                # Handle sensitivity slider
                if sensitivity_box.collidepoint(event.pos):
                    rel_x = event.pos[0] - sensitivity_box.x
                    current_sensitivity = min(5, max(1, int((rel_x / sensitivity_box.width) * 5) + 1))

            if event.type == pygame.KEYDOWN and name_active:
                if event.key == pygame.K_BACKSPACE:
                    user_name = user_name[:-1]
                else:
                    user_name += event.unicode

        # Update colors
        name_color = active_color if name_active else passive_color
        
        # Draw input boxes
        pygame.draw.rect(screen, name_color, name_box, 2)
        draw_button(male_box, 'MALE', BLUE if user_gender == "Male" else passive_color)
        draw_button(female_box, 'FEMALE', (255,105,180) if user_gender == "Female" else passive_color)
        
        # Draw Start button
        if user_name and user_gender:
            draw_button(start_box, 'START GAME', GREEN)
        else:
            draw_button(start_box, 'START GAME', passive_color)

        # Draw Quit Button
        draw_button(quit_box, 'QUIT', RED)
        
        # Draw Sensitivity Slider
        pygame.draw.rect(screen, WHITE, sensitivity_box, 2)
        slider_width = sensitivity_box.width // 5
        for i in range(5):
            color = sensitivity_colors[i] if (i+1) <= current_sensitivity else passive_color
            pygame.draw.rect(screen, color, (sensitivity_box.x + i*slider_width, sensitivity_box.y, slider_width, sensitivity_box.height))
        
        # Draw sensitivity labels
        for i in range(1, 6):
            x_pos = sensitivity_box.x + (i-1)*slider_width + slider_width//2
            draw_text(str(i), info_font, WHITE, screen, x_pos, sensitivity_box.y + 40)

        # Render name text
        name_surface = input_font.render(user_name, True, WHITE)
        screen.blit(name_surface, (name_box.x + 15, name_box.y + 15))
        name_box.w = max(400, name_surface.get_width() + 30)

        pygame.display.update()

# =================================================================
#                           MAIN GAME LOOP
# =================================================================

def game_loop():
    """The main game loop with all gameplay mechanics"""
    global player_x, player_y, player_x_change, player_y_change
    global bullet_x, bullet_y, bullet_state, score_value, lives
    global current_sensitivity

    reset_game_state()
    mixer.music.play(-1)
    
    running = True
    while running:
        screen.blit(background_img, (0, 0))
        
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                save_on_exit()
                running = False
                return
            
            # Keyboard controls for movement
            if event.type == pygame.KEYDOWN:
                if event.key in (pygame.K_LEFT, pygame.K_a):
                    player_x_change = -current_sensitivity
                if event.key in (pygame.K_RIGHT, pygame.K_d):
                    player_x_change = current_sensitivity
                if event.key in (pygame.K_UP, pygame.K_w):
                    player_y_change = -current_sensitivity
                if event.key in (pygame.K_DOWN, pygame.K_s):
                    player_y_change = current_sensitivity
                if event.key == pygame.K_SPACE and bullet_state == "ready":
                    bullet_sound.play()
                    bullet_state = "fire"
                    bullet_x = player_x + 16
                    bullet_y = player_y
            
            if event.type == pygame.KEYUP:
                if event.key in (pygame.K_LEFT, pygame.K_a, pygame.K_RIGHT, pygame.K_d):
                    player_x_change = 0
                if event.key in (pygame.K_UP, pygame.K_w, pygame.K_DOWN, pygame.K_s):
                    player_y_change = 0

        # Update player position with bounds checking
        player_x += player_x_change
        player_y += player_y_change
        player_x = max(0, min(player_x, SCREEN_WIDTH - 64))
        player_y = max(50, min(player_y, SCREEN_HEIGHT - 100))

        # Enemy movement and collision
        for enemy in list(enemies):
            enemy["x"] += enemy["x_change"]
            
            # Enemy bouncing
            if enemy["x"] <= 0 or enemy["x"] >= SCREEN_WIDTH - 64:
                enemy["x_change"] *= -1
                enemy["y"] += 40
            
            # Bullet collision
            if bullet_state == "fire":
                distance = math.sqrt((enemy["x"] - bullet_x)**2 + (enemy["y"] - bullet_y)**2)
                if distance < 27:
                    explosion_sound.play()
                    bullet_state = "ready"
                    score_value += 1
                    enemies.remove(enemy)
                    enemies.append(create_enemy())
            
            # Player collision (same line check)
            if (abs(player_x - enemy["x"]) < 30 and abs(player_y - enemy["y"]) < 30) or enemy["y"] > SCREEN_HEIGHT - 100:
                lives -= 1
                enemies.remove(enemy)
                enemies.append(create_enemy())
                if lives <= 0:
                    running = False

        # Bullet movement
        if bullet_state == "fire":
            bullet_y -= bullet_y_change
            if bullet_y <= 0:
                bullet_state = "ready"

        # Drawing
        if bullet_state == "fire":
            screen.blit(bullet_img, (bullet_x, bullet_y))
        
        screen.blit(player_img, (player_x, player_y))
        
        for enemy in enemies:
            screen.blit(enemy_img, (enemy['x'], enemy['y']))
        
        show_game_ui()
        pygame.display.update()

    # Game over handling
    save_to_excel("Game Over")
    game_over_screen()
    pygame.display.update()
    
    waiting = True
    while waiting:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                exit()
            if event.type == pygame.KEYDOWN and event.key == pygame.K_RETURN:
                waiting = False

# =================================================================
#                      PROGRAM EXECUTION
# =================================================================

if __name__ == "__main__":
    atexit.register(save_on_exit)
    
    while True:
        home_screen()
        game_loop()
        save_on_exit()