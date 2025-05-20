import tkinter as tk
from tkinter import messagebox, ttk 
from random import randint, choice, shuffle
import time
from datetime import datetime
import os 
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

class SudokuGame:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Sudoku Game and Solver")
        self.window.geometry("800x750") 
        self.window.configure(bg='#e0e0e0')

        # User and Game Info
        self.username = tk.StringVar(self.window)
        self.gender = tk.StringVar(self.window)
        self.email = tk.StringVar(self.window)
        self.difficulty = tk.StringVar(self.window, value="easy")
        
        # Game State
        self.hints_remaining = 3
        self.game_mode = None 
        self.start_time = None
        self.game_status = "Not Started" 
        
        self.cells = {} 
        self.solution_grid = [] 
        self.puzzle_grid = []   

        # UI Elements
        self.hints_label = None
        self.hint_button = None 
        self.hint_active = False 
        self.player_entries_before_hint = {}
        self.email_error_label = None 

        self.window.protocol("WM_DELETE_WINDOW", self.handle_quit_game_action)

        self.create_start_menu()

    def create_start_menu(self):
        self.clear_window_widgets()
        self.window.configure(bg='#f0f0f0') # Light gray for start menu
        
        tk.Label(self.window, text="Welcome to Sudoku!", font=('Arial', 30, 'bold'), bg='#f0f0f0', fg='#333').pack(pady=(30,20))
        
        menu_frame = tk.Frame(self.window, bg='#f0f0f0')
        menu_frame.pack(pady=20)

        btn_font = ('Arial', 18)
        # Consistent button options for the main menu
        btn_options = {'font': btn_font, 'fg': 'white', 'padx': 20, 'pady': 10, 'width': 15, 'relief': tk.RAISED, 'bd': 2, 'cursor': 'hand2'}


        tk.Button(menu_frame, text="Play Sudoku", command=self.setup_play_sudoku, bg='#28a745', **btn_options).pack(pady=10)
        tk.Button(menu_frame, text="Solve Sudoku", command=self.setup_solve_sudoku, bg='#007bff', **btn_options).pack(pady=10)
        tk.Button(menu_frame, text="Quit", command=self.handle_quit_game_action, bg='#dc3545', **btn_options).pack(pady=10)


    def setup_play_sudoku(self):
        self.game_mode = "play"
        self.clear_window_widgets()
        self.window.configure(bg='#e9ecef') # Light background for settings

        tk.Label(self.window, text="Player Setup", font=('Arial', 24, 'bold'), bg='#e9ecef', fg='#343a40').pack(pady=20)

        settings_frame = tk.Frame(self.window, bg='#e9ecef')
        settings_frame.pack(pady=10, padx=30, fill=tk.X)
        
        field_font = ('Arial', 14)
        label_font = ('Arial', 14, 'bold')
        btn_cursor = 'hand2'


        tk.Label(settings_frame, text="Username:", font=label_font, bg='#e9ecef').grid(row=0, column=0, padx=5, pady=8, sticky="w")
        tk.Entry(settings_frame, textvariable=self.username, font=field_font, width=25).grid(row=0, column=1, padx=5, pady=8, sticky="ew")

        tk.Label(settings_frame, text="Gender:", font=label_font, bg='#e9ecef').grid(row=1, column=0, padx=5, pady=8, sticky="w")
        gender_frame = tk.Frame(settings_frame, bg='#e9ecef')
        tk.Radiobutton(gender_frame, text="Male", variable=self.gender, value="Male", font=field_font, bg='#e9ecef', cursor=btn_cursor).pack(side=tk.LEFT, padx=5)
        tk.Radiobutton(gender_frame, text="Female", variable=self.gender, value="Female", font=field_font, bg='#e9ecef', cursor=btn_cursor).pack(side=tk.LEFT, padx=5)
        self.gender.set("Male") 
        gender_frame.grid(row=1, column=1, padx=5, pady=8, sticky="w")
        
        tk.Label(settings_frame, text="Email (@gmail.com):", font=label_font, bg='#e9ecef').grid(row=2, column=0, padx=5, pady=8, sticky="w")
        tk.Entry(settings_frame, textvariable=self.email, font=field_font, width=25).grid(row=2, column=1, padx=5, pady=8, sticky="ew")
        self.email_error_label = tk.Label(settings_frame, text="", font=('Arial', 10, 'italic'), fg='red', bg='#e9ecef')
        self.email_error_label.grid(row=3, column=1, padx=5, pady=(0,5), sticky="w")


        tk.Label(settings_frame, text="Select difficulty:", font=label_font, bg='#e9ecef').grid(row=4, column=0, padx=5, pady=8, sticky="w")
        difficulty_menu = tk.OptionMenu(settings_frame, self.difficulty, "easy", "medium", "hard")
        difficulty_menu.config(font=field_font, width=20, relief=tk.GROOVE, bd=2, cursor=btn_cursor)
        difficulty_menu.grid(row=4, column=1, padx=5, pady=8, sticky="ew")
        
        settings_frame.grid_columnconfigure(1, weight=1)

        tk.Button(self.window, text="Start Game", command=self.validate_and_start_play_game, font=('Arial', 18), bg='#28a745', fg='white', padx=15, pady=8, cursor=btn_cursor).pack(pady=(25,10))
        tk.Button(self.window, text="Back to Menu", command=self.create_start_menu, font=('Arial', 14), bg='#6c757d', fg='white', padx=10, pady=5, cursor=btn_cursor).pack(pady=10)

    def validate_and_start_play_game(self):
        username = self.username.get().strip()
        gender = self.gender.get()
        email = self.email.get().strip()

        if not username:
            messagebox.showerror("Input Error", "Username cannot be empty.")
            return
        if not gender: 
            messagebox.showerror("Input Error", "Please select a gender.")
            return
        
        if not email.lower().endswith("@gmail.com"):
            self.email_error_label.config(text="Email must end with @gmail.com")
            # messagebox.showerror("Input Error", "Invalid email format. Must end with @gmail.com.") # Redundant with label
            return
        else:
            self.email_error_label.config(text="")

        self._actual_start_game()

    def _actual_start_game(self):
        self.clear_window_widgets()
        self.hints_remaining = 3 
        self.generate_new_puzzle(self.difficulty.get())
        self.create_sudoku_grid_ui() # This will now correctly set up UI for "play" mode
        self.display_puzzle_on_grid()
        if self.hints_label: self.update_hints_label_text()
        
        self.start_time = time.time() 
        self.game_status = "In Progress"
        print(f"Game started for {self.username.get()} at {time.ctime(self.start_time)}")


    def setup_solve_sudoku(self):
        if self.game_mode == "play" and self.game_status == "In Progress":
            if messagebox.askyesno("Unfinished Game", "You have an unfinished game. Do you want to mark it as 'Quit (Unfinished)' and save your progress before starting the solver?"):
                self._record_game_session("Quit (Unfinished)")
        
        self.game_mode = "solve"
        self.game_status = "Not Started" 
        self.start_time = None # No timer for solver mode
        self.clear_window_widgets()
        self.create_sudoku_grid_ui() # This will now correctly set up UI for "solve" mode
        self.puzzle_grid = [[0 for _ in range(9)] for _ in range(9)]
        self.solution_grid = [[0 for _ in range(9)] for _ in range(9)]
        self.display_puzzle_on_grid() # Displays an empty grid


    def create_sudoku_grid_ui(self):
        self.clear_window_widgets()
        # Use a slightly lighter gray for the game area, similar to the image provided for solver
        self.window.configure(bg='#f0f0f0') 

        main_game_frame = tk.Frame(self.window, bg='#f0f0f0')
        # Pack main_game_frame to allow centering of its children
        main_game_frame.pack(pady=20, padx=20, expand=True, fill=tk.BOTH)


        # --- Header Frame (Only for Play Mode) ---
        if self.game_mode == "play":
            header_frame = tk.Frame(main_game_frame, bg='#f0f0f0')
            header_frame.pack(pady=(0,10), fill=tk.X, anchor=tk.N) # Anchor to North
            if self.username.get():
                tk.Label(header_frame, text=f"Player: {self.username.get()}", font=('Arial', 14, 'italic'), bg='#f0f0f0', fg='#2b3035').pack(side=tk.LEFT, padx=10)
            tk.Label(header_frame, text=f"Difficulty: {self.difficulty.get().capitalize()}", font=('Arial', 14, 'italic'), bg='#f0f0f0', fg='#2b3035').pack(side=tk.LEFT, padx=10)
            self.hints_label = tk.Label(header_frame, text=f"Hints: {self.hints_remaining}", font=('Arial', 14, 'italic'), bg='#f0f0f0', fg='#2b3035')
            self.hints_label.pack(side=tk.LEFT, padx=10)

        # --- Grid Frame (Common for Play and Solve) ---
        # This frame will be centered
        grid_container_frame = tk.Frame(main_game_frame, bg='#f0f0f0')
        grid_container_frame.pack(pady=10, anchor=tk.CENTER) # Center the grid container

        grid_outer_frame = tk.Frame(grid_container_frame, bg='#343a40', bd=3, relief=tk.GROOVE) 
        grid_outer_frame.pack() # Pack normally inside its centered container

        self.cells = {}
        sub_frames = {} 

        for r_block in range(3):
            for c_block in range(3):
                sub_frame = tk.Frame(grid_outer_frame, bg='#adb5bd', bd=1, relief=tk.SUNKEN) 
                sub_frame.grid(row=r_block, column=c_block, padx=1, pady=1)
                sub_frames[(r_block, c_block)] = sub_frame
        
        vcmd = (self.window.register(self.validate_cell_input), '%P', '%S')

        for i in range(9): 
            for j in range(9): 
                block_row, block_col = i // 3, j // 3
                cell_row_in_block, cell_col_in_block = i % 3, j % 3
                parent_frame = sub_frames[(block_row, block_col)]
                
                cell = tk.Entry(parent_frame, width=2, font=('Arial', 22, 'bold'), justify='center',
                                relief=tk.FLAT,
                                highlightbackground='#ced4da', highlightcolor='#007bff', highlightthickness=1,
                                validate="key", validatecommand=vcmd)
                cell.grid(row=cell_row_in_block, column=cell_col_in_block, padx=1, pady=1, ipady=3) 
                self.cells[(i, j)] = cell
                
                cell.bind("<FocusIn>", lambda e, c=cell: c.config(bg='#e6f2ff', highlightthickness=2)) 
                cell.bind("<FocusOut>", lambda e, c=cell: c.config(bg=c.initial_bg_color if hasattr(c, 'initial_bg_color') else '#ffffff', highlightthickness=1))


        # --- Button Frame (Different for Play and Solve) ---
        btn_cursor = 'hand2'
        if self.game_mode == "play":
            play_button_frame = tk.Frame(main_game_frame, bg='#f0f0f0')
            play_button_frame.pack(pady=15, anchor=tk.CENTER) # Center this frame too
            
            btn_font_play = ('Arial', 14)
            btn_options_play = {'font': btn_font_play, 'fg': 'white', 'padx': 10, 'pady': 5, 'width': 10, 'relief': tk.RAISED, 'bd':2, 'cursor': btn_cursor}

            self.hint_button = tk.Button(play_button_frame, text="Hint", command=self.give_hint, bg='#17a2b8', **btn_options_play)
            self.hint_button.pack(side=tk.LEFT, padx=5) 
            tk.Button(play_button_frame, text="Check", command=self.check_player_solution, bg='#28a745', **btn_options_play).pack(side=tk.LEFT, padx=5) 
            tk.Button(play_button_frame, text="New Game", command=self.handle_new_game_request_from_game, bg='#6c757d', **btn_options_play).pack(side=tk.LEFT, padx=5) 
            tk.Button(play_button_frame, text="Quit App", command=self.handle_quit_game_action, bg='#dc3545', **btn_options_play).pack(side=tk.LEFT, padx=5) 

        elif self.game_mode == "solve":
            # Specific layout for solver mode as per the image
            solver_button_frame = tk.Frame(main_game_frame, bg='#f0f0f0')
            solver_button_frame.pack(pady=20, anchor=tk.CENTER) # Center this frame

            # Button styling for solver mode (matching the image)
            solver_btn_font = ('Arial', 16, 'bold') # Slightly larger and bold font
            solver_btn_options = {'font': solver_btn_font, 'fg': 'white', 'padx': 20, 'pady': 8, 'width': 8, 'relief': tk.RAISED, 'bd': 2, 'cursor': btn_cursor}

            # Green color similar to the image for Solve and Clear
            green_button_color = '#4CAF50'
            red_button_color = '#f44336' # A common red, similar to image

            tk.Button(solver_button_frame, text="Solve", command=self.solve_user_puzzle, bg=green_button_color, **solver_btn_options).pack(pady=6) 
            tk.Button(solver_button_frame, text="Clear", command=self.clear_grid_for_solver, bg=green_button_color, **solver_btn_options).pack(pady=6) 
            tk.Button(solver_button_frame, text="Quit", command=self.handle_new_game_request_from_game, bg=red_button_color, **solver_btn_options).pack(pady=6) # "Quit" here should probably go to main menu

    def validate_cell_input(self, P, S): 
        if S.isdigit():
            return len(P) <= 1 and '1' <= S <= '9'
        elif S == "": 
            return True
        else:
            self.window.bell() 
            return False

    def generate_new_puzzle(self, difficulty_str):
        self.solution_grid = [[0 for _ in range(9)] for _ in range(9)]
        self.solve_grid_helper(self.solution_grid, shuffle_numbers=True) 

        self.puzzle_grid = [row[:] for row in self.solution_grid] 

        clues_to_reveal = {"easy": 40, "medium": 32, "hard": 25}.get(difficulty_str, 32)
        cells_to_remove = 81 - clues_to_reveal
        
        removed_count = 0
        all_coords = [(r, c) for r in range(9) for c in range(9)]
        shuffle(all_coords)

        for r, c in all_coords:
            if removed_count >= cells_to_remove:
                break
            if self.puzzle_grid[r][c] != 0:
                self.puzzle_grid[r][c] = 0
                removed_count += 1
        
    def display_puzzle_on_grid(self):
        for i in range(9):
            for j in range(9):
                cell_widget = self.cells[(i, j)]
                cell_widget.config(state='normal') 
                cell_widget.delete(0, tk.END)
                
                value_to_display = self.puzzle_grid[i][j]

                if value_to_display != 0: # Clue cell or pre-filled for solver
                    cell_widget.insert(0, str(value_to_display))
                    if self.game_mode == "play": # Specific styling for play mode clues
                        cell_widget.config(state='disabled', disabledbackground='#e9ecef', disabledforeground='#212529', relief=tk.FLAT)
                        cell_widget.initial_bg_color = '#e9ecef' 
                    else: # Solver mode, allow editing if needed, or style as fixed
                        cell_widget.config(state='normal', bg='#f8f9fa', fg='#495057', relief=tk.FLAT) # Slightly different for solver input
                        cell_widget.initial_bg_color = '#f8f9fa'
                else: # Empty cell for player input or solver input
                    cell_widget.config(state='normal', bg='#ffffff', fg='#0056b3', relief=tk.FLAT) 
                    cell_widget.initial_bg_color = '#ffffff'

    def give_hint(self):
        if self.hint_active:
            messagebox.showinfo("Hint Active", "A hint is currently being displayed. Please wait.")
            return

        if self.game_mode != "play":
            messagebox.showinfo("Hint", "Hints are only available in 'Play Sudoku' mode.")
            return

        if self.hints_remaining <= 0:
            messagebox.showinfo("No Hints", "You have no hints remaining!")
            return

        self.hint_active = True
        if self.hint_button:
            self.hint_button.config(state='disabled')
        
        self.hints_remaining -= 1
        self.update_hints_label_text()

        self.player_entries_before_hint.clear() 
        for r in range(9):
            for c in range(9):
                self.player_entries_before_hint[(r, c)] = self.cells[(r, c)].get()

        for r in range(9):
            for c in range(9):
                cell_widget = self.cells[(r, c)]
                solution_value = str(self.solution_grid[r][c])
                cell_widget.config(state='normal')
                cell_widget.delete(0, tk.END)
                cell_widget.insert(0, solution_value)
                if self.puzzle_grid[r][c] != 0: 
                    cell_widget.config(disabledbackground='#d0d0d0', disabledforeground='#101010', relief=tk.FLAT)
                else: 
                    cell_widget.config(disabledbackground='#cceeff', disabledforeground='#003366', relief=tk.FLAT)
                cell_widget.config(state='disabled')
        
        self.window.after(3000, self._restore_player_view_after_hint)
        print(f"Full solution hint shown. Hints left: {self.hints_remaining}")

    def _restore_player_view_after_hint(self):
        for r in range(9):
            for c in range(9):
                cell_widget = self.cells[(r, c)]
                cell_widget.config(state='normal')
                cell_widget.delete(0, tk.END)
                initial_puzzle_clue = self.puzzle_grid[r][c]
                if initial_puzzle_clue != 0:
                    cell_widget.insert(0, str(initial_puzzle_clue))
                    cell_widget.config(state='disabled', disabledbackground='#e9ecef', disabledforeground='#212529', relief=tk.FLAT)
                    cell_widget.initial_bg_color = '#e9ecef'
                else:
                    player_previous_value = self.player_entries_before_hint.get((r,c), "")
                    if player_previous_value:
                        cell_widget.insert(0, player_previous_value)
                    cell_widget.config(state='normal', bg='#ffffff', fg='#0056b3', relief=tk.FLAT)
                    cell_widget.initial_bg_color = '#ffffff'
                    cell_widget.bind("<FocusIn>", lambda e, cw=cell_widget: cw.config(bg='#e6f2ff', highlightthickness=2))
                    cell_widget.bind("<FocusOut>", lambda e, cw=cell_widget: cw.config(bg=cw.initial_bg_color if hasattr(cw, 'initial_bg_color') else '#ffffff', highlightthickness=1))
        
        self.player_entries_before_hint.clear()
        self.hint_active = False
        if self.hint_button and self.hint_button.winfo_exists(): 
            self.hint_button.config(state='normal')
        print("Player view restored after hint.")


    def update_hints_label_text(self):
        if self.hints_label:
            self.hints_label.config(text=f"Hints: {self.hints_remaining}")

    def check_player_solution(self):
        if self.game_mode != "play": return 

        player_grid = [[0 for _ in range(9)] for _ in range(9)]
        all_filled = True
        valid_numbers = True

        for r in range(9):
            for c in range(9):
                val_str = self.cells[(r, c)].get()
                if not val_str:
                    all_filled = False
                    player_grid[r][c] = 0 
                    continue 
                
                if val_str.isdigit():
                    num = int(val_str)
                    if 1 <= num <= 9:
                        player_grid[r][c] = num
                    else: 
                        valid_numbers = False; break
                else: 
                    valid_numbers = False; break
            if not valid_numbers: break

        if not valid_numbers:
            messagebox.showerror("Invalid Input", "One or more cells contain invalid characters or numbers out of range (1-9).")
            return

        if not all_filled:
            proceed = messagebox.askyesno("Incomplete Puzzle", "The puzzle is not fully filled. Check anyway?")
            if not proceed:
                return

        correct_so_far = True
        for r_idx in range(9):
            for c_idx in range(9):
                cell_widget = self.cells[(r_idx,c_idx)]
                if cell_widget.cget('state') == 'normal': 
                     cell_widget.config(bg=cell_widget.initial_bg_color if hasattr(cell_widget, 'initial_bg_color') else '#ffffff', 
                                        fg='#0056b3')

        for r in range(9):
            for c in range(9):
                if player_grid[r][c] != 0 and player_grid[r][c] != self.solution_grid[r][c]:
                    correct_so_far = False
                    self.cells[(r,c)].config(bg='#f8d7da', fg='#721c24') 
                elif player_grid[r][c] != 0 and player_grid[r][c] == self.solution_grid[r][c] and self.puzzle_grid[r][c] == 0 : 
                     self.cells[(r,c)].config(bg='#d4edda', fg='#155724') 
        
        if correct_so_far and all_filled: 
            self.game_status = "Solved"
            self._record_game_session(self.game_status)
            messagebox.showinfo("Congratulations!", f"{self.username.get()}, you solved the Sudoku correctly!")
        elif correct_so_far and not all_filled:
            messagebox.showinfo("Good Progress!", "All filled numbers are correct so far. Keep going!")
        else: 
             messagebox.showerror("Incorrect", "Some numbers are incorrect. They have been highlighted.")


    def solve_user_puzzle(self): 
        user_puzzle_data = [[0 for _ in range(9)] for _ in range(9)]
        input_is_valid = True
        has_input = False # Check if user entered any numbers
        for r in range(9):
            for c in range(9):
                val_str = self.cells[(r, c)].get()
                if val_str:
                    has_input = True
                    if val_str.isdigit() and 1 <= int(val_str) <= 9:
                        user_puzzle_data[r][c] = int(val_str)
                    else:
                        messagebox.showerror("Invalid Input", f"Invalid character '{val_str}' or number out of range at ({r+1},{c+1}).", parent=self.window)
                        input_is_valid = False; break
                else:
                    user_puzzle_data[r][c] = 0
            if not input_is_valid: return
        
        if not input_is_valid: return
        if not has_input:
            messagebox.showinfo("Empty Grid", "Please enter some numbers into the grid to solve.", parent=self.window)
            return

        if not self.is_initial_grid_valid(user_puzzle_data):
            messagebox.showerror("Invalid Puzzle", "The initial numbers in the puzzle conflict with Sudoku rules.", parent=self.window)
            return

        # Create a copy to solve, so the original UI entries are preserved if no solution
        solve_attempt_grid = [row[:] for row in user_puzzle_data]

        if self.solve_grid_helper(solve_attempt_grid, shuffle_numbers=False):
            self.solution_grid = [row[:] for row in solve_attempt_grid] 
            self.puzzle_grid = [row[:] for row in solve_attempt_grid] 
            # Update UI with the solved grid
            for r_idx in range(9):
                for c_idx in range(9):
                    cell_widget = self.cells[(r_idx, c_idx)]
                    cell_widget.config(state='normal')
                    cell_widget.delete(0, tk.END)
                    cell_widget.insert(0, str(self.solution_grid[r_idx][c_idx]))
                    # Style solved cells, differentiate from original input if desired
                    if user_puzzle_data[r_idx][c_idx] == 0: # Cell was originally empty, now filled by solver
                        cell_widget.config(state='disabled', disabledbackground='#d4edda', disabledforeground='#155724') # Greenish
                    else: # Cell was part of user's initial input
                        cell_widget.config(state='disabled', disabledbackground='#e9ecef', disabledforeground='#212529') # Original clue style
            messagebox.showinfo("Solved!", "The Sudoku has been solved.", parent=self.window)
        else:
            messagebox.showerror("No Solution", "This Sudoku puzzle has no solution or is invalid based on the input.", parent=self.window)

    def clear_grid_for_solver(self): 
        for r in range(9):
            for c in range(9):
                cell = self.cells[(r, c)]
                cell.config(state='normal')
                cell.delete(0, tk.END)
                cell.config(bg='#ffffff', fg='#0056b3', relief=tk.FLAT) 
                cell.initial_bg_color = '#ffffff'
        self.puzzle_grid = [[0 for _ in range(9)] for _ in range(9)] 
        self.solution_grid = [[0 for _ in range(9)] for _ in range(9)]


    def find_empty_cell(self, grid):
        for r in range(9):
            for c in range(9):
                if grid[r][c] == 0:
                    return (r, c)
        return None

    def is_move_valid(self, grid, num, pos): 
        row, col = pos
        for c_idx in range(9):
            if grid[row][c_idx] == num and c_idx != col: 
                return False
        for r_idx in range(9):
            if grid[r_idx][col] == num and r_idx != row: 
                return False
        box_start_row, box_start_col = row - row % 3, col - col % 3
        for r_idx in range(box_start_row, box_start_row + 3):
            for c_idx in range(box_start_col, box_start_col + 3):
                if grid[r_idx][c_idx] == num and (r_idx != row or c_idx != col): 
                    return False
        return True
        
    def is_initial_grid_valid(self, grid): 
        for r in range(9):
            for c in range(9):
                num = grid[r][c]
                if num != 0: 
                    grid[r][c] = 0 
                    if not self.is_move_valid(grid, num, (r,c)): 
                        grid[r][c] = num 
                        return False
                    grid[r][c] = num 
        return True

    def solve_grid_helper(self, grid_to_solve, shuffle_numbers=False):
        empty_pos = self.find_empty_cell(grid_to_solve)
        if not empty_pos:
            return True 

        row, col = empty_pos
        numbers_to_try = list(range(1, 10))
        if shuffle_numbers:
            shuffle(numbers_to_try)

        for num in numbers_to_try:
            if self.is_move_valid(grid_to_solve, num, (row, col)): 
                grid_to_solve[row][col] = num 
                if self.solve_grid_helper(grid_to_solve, shuffle_numbers): 
                    return True
                grid_to_solve[row][col] = 0 
        return False 


    def clear_window_widgets(self):
        for widget in self.window.winfo_children():
            widget.destroy()
        self.cells = {}
        self.hints_label = None
        self.hint_button = None
        self.hint_active = False
        self.email_error_label = None 

    def _format_time(self, seconds_elapsed):
        if seconds_elapsed is None:
            return "N/A"
        s = int(seconds_elapsed)
        hours = s // 3600
        s %= 3600
        minutes = s // 60
        seconds = s % 60
        return f"{hours:02}:{minutes:02}:{seconds:02}"

    def _record_game_session(self, final_status):
        if self.game_mode != "play" or not self.username.get() or self.start_time is None:
            print("Condition not met for recording game session.")
            return

        elapsed_seconds = time.time() - self.start_time
        formatted_time = self._format_time(elapsed_seconds)
        self.game_status = final_status 

        data_to_save = {
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Username": self.username.get(),
            "Gender": self.gender.get(),
            "Email": self.email.get(),
            "Difficulty": self.difficulty.get(),
            "Time Taken (s)": round(elapsed_seconds, 2),
            "Time Taken (HH:MM:SS)": formatted_time,
            "Status": self.game_status
        }
        
        self.save_game_data_to_excel(data_to_save)
        
        self.start_time = None 

    def save_game_data_to_excel(self, data):
        username = data["Username"]
        filename = f"Sudoku of {username}.xlsx"
        
        headers = ["Timestamp", "Username", "Gender", "Email", "Difficulty", "Time Taken (s)", "Time Taken (HH:MM:SS)", "Status"]
        row_data = [data[h] for h in headers] 

        try:
            if os.path.exists(filename):
                try:
                    workbook = load_workbook(filename)
                    sheet = workbook.active
                except InvalidFileException: 
                    print(f"Warning: File '{filename}' is not a valid Excel file or is corrupted. Creating a new one.")
                    workbook = Workbook()
                    sheet = workbook.active
                    sheet.append(headers) 
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(headers) 
            
            sheet.append(row_data)
            workbook.save(filename)
            print(f"Data saved to {filename}")
            messagebox.showinfo("Data Saved", f"Your game data has been saved to {filename}", parent=self.window)

        except Exception as e:
            print(f"Error saving data to Excel: {e}")
            messagebox.showerror("Excel Error", f"Could not save data to Excel: {e}", parent=self.window)

    def handle_new_game_request_from_game(self):
        if self.game_mode == "play" and self.game_status == "In Progress":
            if messagebox.askyesno("Unfinished Game", "You have an unfinished game. Do you want to mark it as 'Quit (Unfinished)' and save your progress before starting a new game/session?", parent=self.window):
                self._record_game_session("Quit (Unfinished)")
        elif self.game_mode == "solve": # If quitting from solver mode
            pass # No game data to save for solver mode, just go to menu

        self.start_time = None
        self.game_status = "Not Started"
        self.create_start_menu() 

    def handle_quit_game_action(self):
        quit_confirmed = False
        if self.game_mode == "play" and self.game_status == "In Progress":
            if messagebox.askyesno("Quit Game", "You have an unfinished game. Do you want to mark it as 'Quit (Unsolved)' and save your progress before quitting?", parent=self.window):
                self._record_game_session("Quit (Unsolved)")
                quit_confirmed = True 
            elif messagebox.askyesno("Confirm Quit", "Are you sure you want to quit without saving the current game?", parent=self.window):
                quit_confirmed = True 
        else:
            if messagebox.askyesno("Quit Application", "Are you sure you want to quit Sudoku?", parent=self.window):
                quit_confirmed = True
        
        if quit_confirmed:
            self.window.destroy() 

    def run(self):
        self.window.mainloop()

if __name__ == "__main__":
    game = SudokuGame()
    game.run()
