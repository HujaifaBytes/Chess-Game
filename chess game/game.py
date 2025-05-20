import tkinter as tk
from tkinter import messagebox, font
import time
import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font # Import Font for styling
from PIL import Image, ImageTk
from openpyxl.utils import get_column_letter

class ChessGame:
    def __init__(self, root):
        self.root = root
        self.root.title("Two-Player Chess Game")
        self.root.geometry("900x750")
        self.root.configure(bg="#333333")

        # Constants
        self.BOARD_SIZE = 8
        self.SQUARE_SIZE = 70
        self.INFO_PANEL_WIDTH = 200
        self.BUTTON_WIDTH = 10
        self.BUTTON_HEIGHT = 2

        # Colors
        self.LIGHT_SQUARE = "#D3D3D3"
        self.DARK_SQUARE = "#808080"
        self.HIGHLIGHT_WHITE = "#FF0000"
        self.HIGHLIGHT_BLACK = "#0000FF"
        self.HIGHLIGHT_CHECK_WHITE = "#8B0000" # Dark Red for White King in Check
        self.HIGHLIGHT_CHECK_BLACK = "#00008B" # Dark Blue for Black King in Check
        self.VALID_MOVE_COLOR = "#00FF00"
        self.PANEL_COLOR = "#4A4A4A"
        self.BORDER_COLOR = "#FFD700"
        self.RESTART_COLOR = "#ADD8E6"
        self.QUIT_COLOR = "#FF4500"

        # Fonts
        self.status_font = font.Font(family="Arial", size=14, weight="bold")
        self.button_font = font.Font(family="Arial", size=12, weight="bold")

        # Game variables
        self.player_white_name = ""
        self.player_black_name = ""
        self.game_start_time = None
        self.total_moves_count = 0
        # self.counter = 0 # Removed, no longer needed
        self.white_pieces = []
        self.white_locations = []
        self.black_pieces = []
        self.black_locations = []
        self.captured_pieces_white = [] # Stores names of pieces captured by white
        self.captured_pieces_black = [] # Stores names of pieces captured by black
        self.turn_step = 0 # 0,1 for white; 2,3 for black
        self.selection = 100 # Index of selected piece, 100 for none
        self.valid_moves = []
        self.winner = '' # 'white' or 'black'
        self.game_over = False
        self.white_options = []
        self.black_options = []
        self.piece_list = ['pawn', 'queen', 'king', 'knight', 'rook', 'bishop'] # Keep for reference
        self.white_promotions = ['bishop', 'knight', 'rook', 'queen'] # For future promotion logic
        self.black_promotions = ['bishop', 'knight', 'rook', 'queen'] # For future promotion logic
        self.game_end_condition = "Unknown"

        # Load images
        self.image_dict = self.load_images()

        # UI setup
        self.show_home_screen()

    def load_images(self):
        image_dict = {}
        base_path = r"D:\chess game" # Consider making this a relative path
        piece_files = {
            'white': {
                'pawn': 'white_pawn.png', 'queen': 'white_queen.png', 'king': 'white_king.png',
                'knight': 'white_knight.png', 'rook': 'white_rook.png', 'bishop': 'white_bishop.png'
            },
            'black': {
                'pawn': 'black_pawn.png', 'queen': 'black_queen.png', 'king': 'black_king.png',
                'knight': 'black_knight.png', 'rook': 'black_rook.png', 'bishop': 'black_bishop.png'
            }
        }
        for color in ['white', 'black']:
            for piece in self.piece_list: # Uses self.piece_list correctly
                file_path = os.path.join(base_path, piece_files[color][piece])
                try:
                    img = Image.open(file_path)
                    img = img.resize((self.SQUARE_SIZE - 10, self.SQUARE_SIZE - 10), Image.Resampling.LANCZOS)
                    image_dict[f"{color}_{piece}"] = ImageTk.PhotoImage(img)
                except FileNotFoundError as e:
                    messagebox.showerror("Image Load Error", f"Image not found: {file_path}\nPlease ensure images are in '{base_path}'.")
                    print(f"Image not found: {file_path}")
                    # self.root.destroy() # Optionally exit if images are critical
                    raise # Or handle more gracefully
                except Exception as e:
                    messagebox.showerror("Image Load Error", f"Error loading {file_path}: {e}")
                    raise
        return image_dict

    def show_home_screen(self):
        self.clear_screen()
        # Reset any lingering game state when returning to home
        self.game_start_time = None
        self.winner = ''
        self.game_over = False
        self.game_end_condition = "Unknown"

        self.home_frame = tk.Frame(self.root, bg="#333333")
        self.home_frame.pack(expand=True)

        tk.Label(self.home_frame, text="Enter Player Names", bg="#333333", fg="white",
                 font=font.Font(family="Arial", size=18, weight="bold")).pack(pady=20)

        tk.Label(self.home_frame, text="Player 1 (White):", bg="#333333", fg="white",
                 font=self.status_font).pack(pady=5)
        self.white_entry = tk.Entry(self.home_frame, font=self.status_font)
        self.white_entry.pack(pady=5)

        tk.Label(self.home_frame, text="Player 2 (Black):", bg="#333333", fg="white",
                 font=self.status_font).pack(pady=5)
        self.black_entry = tk.Entry(self.home_frame, font=self.status_font)
        self.black_entry.pack(pady=5)

        tk.Button(self.home_frame, text="Start Game", command=self.start_game,
                  bg="#ADD8E6", fg="black", font=self.button_font,
                  width=self.BUTTON_WIDTH + 5, height=self.BUTTON_HEIGHT-1).pack(pady=20)

        self.quit_button_home = tk.Button(self.home_frame, text="Quit", command=self.quit_app,
                                     bg=self.QUIT_COLOR, fg="black", font=self.button_font,
                                     width=self.BUTTON_WIDTH, height=self.BUTTON_HEIGHT)
        self.quit_button_home.pack(pady=5)

    def clear_screen(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def start_game(self):
        white_name = self.white_entry.get().strip()
        black_name = self.black_entry.get().strip()
        if white_name and black_name:
            self.player_white_name = white_name
            self.player_black_name = black_name
            self.clear_screen()
            self.setup_ui()
            self.setup_new_game() # This will set game_start_time
            self.update_ui()
        else:
            messagebox.showerror("Error", "Please enter names for both players.")

    def setup_ui(self):
        self.main_frame = tk.Frame(self.root, bg="#333333")
        self.main_frame.pack(pady=10)

        self.board_canvas = tk.Canvas(self.main_frame, width=self.BOARD_SIZE * self.SQUARE_SIZE,
                                      height=self.BOARD_SIZE * self.SQUARE_SIZE, bg="#333333", highlightthickness=0)
        self.board_canvas.grid(row=0, column=0, padx=(0, 10))

        self.side_panel = tk.Frame(self.main_frame, width=self.INFO_PANEL_WIDTH,
                                   height=self.BOARD_SIZE * self.SQUARE_SIZE, bg=self.PANEL_COLOR,
                                   highlightbackground=self.BORDER_COLOR, highlightthickness=5)
        self.side_panel.grid(row=0, column=1, sticky="ns")
        self.side_panel.grid_propagate(False)

        self.captured_white_label = tk.Label(self.side_panel, text="Captured by White:", bg=self.PANEL_COLOR,
                                            fg="white", font=self.status_font)
        self.captured_white_label.grid(row=0, column=0, pady=(10, 5), padx=5, sticky="w")
        self.captured_white_pieces = tk.Label(self.side_panel, text="", bg=self.PANEL_COLOR,
                                             fg="white", font=self.status_font, wraplength=self.INFO_PANEL_WIDTH-20)
        self.captured_white_pieces.grid(row=1, column=0, padx=5, sticky="w")

        self.captured_black_label = tk.Label(self.side_panel, text="Captured by Black:", bg=self.PANEL_COLOR,
                                            fg="white", font=self.status_font)
        self.captured_black_label.grid(row=2, column=0, pady=(20, 5), padx=5, sticky="w")
        self.captured_black_pieces = tk.Label(self.side_panel, text="", bg=self.PANEL_COLOR,
                                             fg="white", font=self.status_font, wraplength=self.INFO_PANEL_WIDTH-20)
        self.captured_black_pieces.grid(row=3, column=0, padx=5, sticky="w")

        self.bottom_panel = tk.Frame(self.root, width=self.BOARD_SIZE * self.SQUARE_SIZE + self.INFO_PANEL_WIDTH + 10,
                                     height=100, bg=self.PANEL_COLOR, highlightbackground=self.BORDER_COLOR,
                                     highlightthickness=5)
        self.bottom_panel.pack(pady=(10, 0))
        self.bottom_panel.pack_propagate(False)

        self.turn_label = tk.Label(self.bottom_panel, text="White Turn", bg=self.PANEL_COLOR,
                                   fg="white", font=self.status_font)
        self.turn_label.pack(side=tk.LEFT, padx=20, pady=10)

        self.button_frame = tk.Frame(self.bottom_panel, bg=self.PANEL_COLOR)
        self.button_frame.pack(side=tk.RIGHT, padx=20)

        self.quit_button_game = tk.Button(self.button_frame, text="QUIT", command=self.return_to_home, # Changed from self.quit_app
                                     bg=self.QUIT_COLOR, fg="black", font=self.button_font,
                                     width=self.BUTTON_WIDTH, height=self.BUTTON_HEIGHT)
        self.quit_button_game.pack(side=tk.RIGHT, padx=(5, 0)) # Swapped order for common UI pattern

        self.restart_button = tk.Button(self.button_frame, text="RESTART", command=self.restart_game,
                                        bg=self.RESTART_COLOR, fg="black", font=self.button_font,
                                        width=self.BUTTON_WIDTH, height=self.BUTTON_HEIGHT)
        self.restart_button.pack(side=tk.RIGHT, padx=(0,5))


        self.board_canvas.bind("<Button-1>", self.handle_click)

    def setup_new_game(self):
        self.white_pieces = ['rook', 'knight', 'bishop', 'king', 'queen', 'bishop', 'knight', 'rook',
                             'pawn', 'pawn', 'pawn', 'pawn', 'pawn', 'pawn', 'pawn', 'pawn']
        self.white_locations = [(0, 0), (1, 0), (2, 0), (3, 0), (4, 0), (5, 0), (6, 0), (7, 0),
                                (0, 1), (1, 1), (2, 1), (3, 1), (4, 1), (5, 1), (6, 1), (7, 1)]
        self.black_pieces = ['rook', 'knight', 'bishop', 'king', 'queen', 'bishop', 'knight', 'rook',
                             'pawn', 'pawn', 'pawn', 'pawn', 'pawn', 'pawn', 'pawn', 'pawn']
        self.black_locations = [(0, 7), (1, 7), (2, 7), (3, 7), (4, 7), (5, 7), (6, 7), (7, 7),
                                (0, 6), (1, 6), (2, 6), (3, 6), (4, 6), (5, 6), (6, 6), (7, 6)]
        self.captured_pieces_white = []
        self.captured_pieces_black = []
        self.turn_step = 0
        self.selection = 100
        self.valid_moves = []
        self.winner = ''
        self.game_over = False
        self.game_start_time = time.time() # Game starts now
        self.total_moves_count = 0
        self.game_end_condition = "Ongoing" # Or "Not Started Yet" if preferred
        # Recalculate options for the new board setup
        self.black_options = self.check_options(self.black_pieces, self.black_locations, 'black')
        self.white_options = self.check_options(self.white_pieces, self.white_locations, 'white')
        # self.update_ui() # update_ui is usually called after setup_new_game by the caller

    def draw_board(self):
        self.board_canvas.delete("all") # Clear previous drawings
        for row in range(self.BOARD_SIZE):
            for col in range(self.BOARD_SIZE):
                x1 = col * self.SQUARE_SIZE
                y1 = row * self.SQUARE_SIZE
                x2 = x1 + self.SQUARE_SIZE
                y2 = y1 + self.SQUARE_SIZE
                color = self.LIGHT_SQUARE if (row + col) % 2 == 0 else self.DARK_SQUARE
                self.board_canvas.create_rectangle(x1, y1, x2, y2, fill=color, outline="black")

    def draw_pieces(self):
        # Delete only piece-specific tags, not "all" which is done by draw_board
        self.board_canvas.delete("piece", "highlight", "valid", "check")

        for i, (piece, loc) in enumerate(zip(self.white_pieces, self.white_locations)):
            x_center = loc[0] * self.SQUARE_SIZE + self.SQUARE_SIZE // 2
            y_center = loc[1] * self.SQUARE_SIZE + self.SQUARE_SIZE // 2
            self.board_canvas.create_image(x_center, y_center, image=self.image_dict[f"white_{piece}"], tags="piece")
            if self.turn_step < 2 and self.selection == i: # White's turn and this piece is selected
                self.board_canvas.create_rectangle(loc[0] * self.SQUARE_SIZE + 2, loc[1] * self.SQUARE_SIZE + 2,
                                                   (loc[0] + 1) * self.SQUARE_SIZE - 2, (loc[1] + 1) * self.SQUARE_SIZE - 2,
                                                   outline=self.HIGHLIGHT_WHITE, width=3, tags="highlight")

        for i, (piece, loc) in enumerate(zip(self.black_pieces, self.black_locations)):
            x_center = loc[0] * self.SQUARE_SIZE + self.SQUARE_SIZE // 2
            y_center = loc[1] * self.SQUARE_SIZE + self.SQUARE_SIZE // 2
            self.board_canvas.create_image(x_center, y_center, image=self.image_dict[f"black_{piece}"], tags="piece")
            if self.turn_step >= 2 and self.selection == i: # Black's turn and this piece is selected
                self.board_canvas.create_rectangle(loc[0] * self.SQUARE_SIZE + 2, loc[1] * self.SQUARE_SIZE + 2,
                                                   (loc[0] + 1) * self.SQUARE_SIZE - 2, (loc[1] + 1) * self.SQUARE_SIZE - 2,
                                                   outline=self.HIGHLIGHT_BLACK, width=3, tags="highlight")
        
        if self.selection != 100 and not self.game_over: # Only show valid moves if a piece is selected and game not over
            # self.valid_moves = self.check_valid_moves() # Already calculated or should be updated when selection changes
            for move in self.valid_moves:
                x_center = move[0] * self.SQUARE_SIZE + self.SQUARE_SIZE // 2
                y_center = move[1] * self.SQUARE_SIZE + self.SQUARE_SIZE // 2
                self.board_canvas.create_oval(x_center - 5, y_center - 5, x_center + 5, y_center + 5,
                                              fill=self.VALID_MOVE_COLOR, tags="valid")
        self.draw_check() # Draw check status

    def draw_captured(self):
        white_captured_display = " ".join([p[0].upper() for p in self.captured_pieces_white])
        black_captured_display = " ".join([p[0].upper() for p in self.captured_pieces_black])
        self.captured_white_pieces.config(text=white_captured_display)
        self.captured_black_pieces.config(text=black_captured_display)

    def draw_check(self):
        self.board_canvas.delete("check") # Clear previous check highlights
        # Show check for White King (if it's White's turn or for general display)
        if 'king' in self.white_pieces:
            king_idx = self.white_pieces.index('king')
            king_loc = self.white_locations[king_idx]
            for options_list in self.black_options: # black_options are lists of moves for each black piece
                if king_loc in options_list:
                    self.board_canvas.create_rectangle(king_loc[0] * self.SQUARE_SIZE + 2, king_loc[1] * self.SQUARE_SIZE + 2,
                                                       (king_loc[0] + 1) * self.SQUARE_SIZE - 2, (king_loc[1] + 1) * self.SQUARE_SIZE - 2,
                                                       outline=self.HIGHLIGHT_CHECK_WHITE, width=3, tags="check")
                    break # King is in check, no need to check other black pieces
        
        # Show check for Black King
        if 'king' in self.black_pieces:
            king_idx = self.black_pieces.index('king')
            king_loc = self.black_locations[king_idx]
            for options_list in self.white_options:
                if king_loc in options_list:
                    self.board_canvas.create_rectangle(king_loc[0] * self.SQUARE_SIZE + 2, king_loc[1] * self.SQUARE_SIZE + 2,
                                                       (king_loc[0] + 1) * self.SQUARE_SIZE - 2, (king_loc[1] + 1) * self.SQUARE_SIZE - 2,
                                                       outline=self.HIGHLIGHT_CHECK_BLACK, width=3, tags="check")
                    break

    def process_game_over_prompt(self):
        winner_name = ""
        if self.winner == 'white':
            winner_name = self.player_white_name
        elif self.winner == 'black':
            winner_name = self.player_black_name
        else: # Should not happen if self.winner is set
            winner_name = "Nobody"

        response = messagebox.askyesno("Game Over", f"{winner_name} won!\nDo you want to play another game?")

        if response:  # User clicked "Yes"
            self.game_end_condition = "Finished"
            self.save_game_to_excel()
            self.setup_new_game()  # Resets game_over, winner, etc.
            self.update_ui()
        else:  # User clicked "No"
            # Do not save this game's "Finished" state.
            # Return to home screen without saving current game outcome.
            self.return_to_home(from_game_end_decline=True)

    def update_ui(self):
        if not hasattr(self, 'board_canvas') or not self.board_canvas.winfo_exists():
             return # UI not ready or destroyed

        self.draw_board() # Redraws squares
        self.draw_pieces() # Redraws pieces, highlights, valid moves, and check
        self.draw_captured()
        current_turn_text = ""
        if not self.game_over:
            current_turn_text = f"{self.player_white_name}'s Turn (White)" if self.turn_step < 2 else f"{self.player_black_name}'s Turn (Black)"
        elif self.winner:
             winner_player = self.player_white_name if self.winner == 'white' else self.player_black_name
             current_turn_text = f"Game Over! {winner_player} won."
        else:
             current_turn_text = "Game Over!" # e.g. if draw condition met in future
        self.turn_label.config(text=current_turn_text)
        self.root.update_idletasks() # Process pending Tkinter operations
        # self.root.update() # Can sometimes cause issues or be unnecessary

    def check_options(self, pieces, locations, color):
        all_moves_list = []
        for i in range(len(pieces)):
            location = locations[i]
            piece = pieces[i]
            current_piece_moves = []
            if piece == 'pawn':
                current_piece_moves = self.check_pawn(location, color)
            elif piece == 'rook':
                current_piece_moves = self.check_rook(location, color)
            elif piece == 'knight':
                current_piece_moves = self.check_knight(location, color)
            elif piece == 'bishop':
                current_piece_moves = self.check_bishop(location, color)
            elif piece == 'queen':
                current_piece_moves = self.check_queen(location, color)
            elif piece == 'king':
                current_piece_moves = self.check_king(location, color)
            all_moves_list.append(current_piece_moves)
        return all_moves_list

    def get_enemies_friends(self, color):
        if color == 'white':
            friends_list = self.white_locations
            enemies_list = self.black_locations
        else: # color == 'black'
            friends_list = self.black_locations
            enemies_list = self.white_locations
        return friends_list, enemies_list

    def check_king(self, position, color):
        moves_list = []
        friends_list, _ = self.get_enemies_friends(color) # King cannot move to a square occupied by a friendly piece
        # King moves one square in any direction
        for dr in [-1, 0, 1]:
            for dc in [-1, 0, 1]:
                if dr == 0 and dc == 0:
                    continue # Skip current position
                target_row, target_col = position[1] + dr, position[0] + dc # row is y (pos[1]), col is x (pos[0])
                if 0 <= target_row <= 7 and 0 <= target_col <= 7:
                    target_pos = (target_col, target_row)
                    if target_pos not in friends_list:
                        moves_list.append(target_pos)
        return moves_list

    def check_queen(self, position, color):
        moves_list = self.check_bishop(position, color)
        moves_list.extend(self.check_rook(position, color))
        return moves_list

    def check_bishop(self, position, color):
        moves_list = []
        friends_list, enemies_list = self.get_enemies_friends(color)
        # Directions: diagonal
        for dr, dc in [(-1, -1), (-1, 1), (1, -1), (1, 1)]:
            for i in range(1, 8): # Max 7 steps
                target_row, target_col = position[1] + i * dr, position[0] + i * dc
                if 0 <= target_row <= 7 and 0 <= target_col <= 7:
                    target_pos = (target_col, target_row)
                    if target_pos in friends_list: # Blocked by a friendly piece
                        break
                    moves_list.append(target_pos)
                    if target_pos in enemies_list: # Can capture enemy, but cannot move further in this direction
                        break
                else: # Off board
                    break
        return moves_list

    def check_rook(self, position, color):
        moves_list = []
        friends_list, enemies_list = self.get_enemies_friends(color)
        # Directions: horizontal and vertical
        for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
            for i in range(1, 8):
                target_row, target_col = position[1] + i * dr, position[0] + i * dc
                if 0 <= target_row <= 7 and 0 <= target_col <= 7:
                    target_pos = (target_col, target_row)
                    if target_pos in friends_list:
                        break
                    moves_list.append(target_pos)
                    if target_pos in enemies_list:
                        break
                else:
                    break
        return moves_list

    def check_pawn(self, position, color):
        moves_list = []
        # friends_list, enemies_list = self.get_enemies_friends(color) # Not directly needed here, using white/black locations
        r, c = position[1], position[0] # row, col for clarity

        if color == 'white':
            direction = 1 # White pawns move from row 1 towards row 7
            start_row = 1
            # Single step forward
            if r + direction <= 7 and (c, r + direction) not in self.white_locations and (c, r + direction) not in self.black_locations:
                moves_list.append((c, r + direction))
                # Double step from start
                if r == start_row and (c, r + 2 * direction) not in self.white_locations and (c, r + 2 * direction) not in self.black_locations:
                    moves_list.append((c, r + 2 * direction))
            # Captures
            if c - 1 >= 0 and r + direction <= 7 and (c - 1, r + direction) in self.black_locations:
                moves_list.append((c - 1, r + direction))
            if c + 1 <= 7 and r + direction <= 7 and (c + 1, r + direction) in self.black_locations:
                moves_list.append((c + 1, r + direction))
        else: # color == 'black'
            direction = -1 # Black pawns move from row 6 towards row 0
            start_row = 6
            # Single step forward
            if r + direction >= 0 and (c, r + direction) not in self.black_locations and (c, r + direction) not in self.white_locations:
                moves_list.append((c, r + direction))
                # Double step from start
                if r == start_row and (c, r + 2 * direction) not in self.black_locations and (c, r + 2 * direction) not in self.white_locations:
                    moves_list.append((c, r + 2 * direction))
            # Captures
            if c - 1 >= 0 and r + direction >= 0 and (c - 1, r + direction) in self.white_locations:
                moves_list.append((c - 1, r + direction))
            if c + 1 <= 7 and r + direction >= 0 and (c + 1, r + direction) in self.white_locations:
                moves_list.append((c + 1, r + direction))
        return moves_list
    
    def check_knight(self, position, color):
        moves_list = []
        friends_list, _ = self.get_enemies_friends(color)
        r, c = position[1], position[0]
        # Possible L-shaped moves for a knight
        knight_moves = [(-2, -1), (-2, 1), (-1, -2), (-1, 2),
                        (1, -2), (1, 2), (2, -1), (2, 1)]
        for dr, dc in knight_moves:
            target_row, target_col = r + dr, c + dc
            if 0 <= target_row <= 7 and 0 <= target_col <= 7:
                target_pos = (target_col, target_row)
                if target_pos not in friends_list:
                    moves_list.append(target_pos)
        return moves_list

    def check_valid_moves(self): # Get valid moves for the currently selected piece
        if self.selection == 100 or self.game_over: # No piece selected or game is over
            return []

        if self.turn_step < 2: # White's turn
            if self.selection < len(self.white_options):
                return self.white_options[self.selection]
        else: # Black's turn
            if self.selection < len(self.black_options):
                return self.black_options[self.selection]
        return []


    def format_duration(self, seconds):
        s = int(seconds)
        hours = s // 3600
        s %= 3600
        minutes = s // 60
        secs = s % 60
        return f"{hours:02}:{minutes:02}:{secs:02}"
    def _generate_game_log_filename(self, base_name="Chess_Game_Log"):
            assets_dir = 'assets'
            if not os.path.exists(assets_dir):
                os.makedirs(assets_dir)
            
            i = 1
            while True:
                filename = os.path.join(assets_dir, f"{base_name}_{i}.xlsx")
                if not os.path.exists(filename):
                    return filename
                i += 1
    def save_game_to_excel(self):
        if self.game_start_time is None:
            print("No game data to save (game not started or already reset).")
            return

        duration_seconds = time.time() - self.game_start_time
        duration_formatted = self.format_duration(duration_seconds)

        actual_winner_name = "Draw"
        if self.winner == 'white':
            actual_winner_name = self.player_white_name
        elif self.winner == 'black':
            actual_winner_name = self.player_black_name
        
        if self.game_end_condition in ["Quit", "Reset"]:
            actual_winner_name = "N/A"

        game_data = {
            "Timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "PlayerWhite": self.player_white_name,
            "PlayerBlack": self.player_black_name,
            "Winner": actual_winner_name,
            "DurationSeconds": round(duration_seconds, 2),
            "DurationFormatted": duration_formatted,
            "TotalMoves": self.total_moves_count,
            "GameStatus": self.game_end_condition
        }

        assets_dir = 'assets'
        if not os.path.exists(assets_dir):
            os.makedirs(assets_dir)

        filename = os.path.join(assets_dir, "Chess_data.xlsx")
        headers = ["Timestamp", "PlayerWhite", "PlayerBlack", "Winner",
                   "DurationSeconds", "DurationFormatted", "TotalMoves", "GameStatus"]
        
        workbook = None
        sheet = None

        try:
            if os.path.exists(filename):
                workbook = load_workbook(filename)
                if "GameLog" in workbook.sheetnames:
                    sheet = workbook["GameLog"]
                    print(f"Loaded existing 'GameLog' sheet with {sheet.max_row} data rows (excluding header).")
                else:
                    # If "GameLog" sheet is missing in an existing file, create it.
                    print(f"Warning: 'GameLog' sheet not found in {filename}. Creating it.")
                    sheet = workbook.create_sheet("GameLog")
                    sheet.append(headers)
                    for cell_h in sheet[1]: # First row (headers)
                        cell_h.font = Font(bold=True)
            else: # File does not exist, create new workbook and sheet
                workbook = Workbook()
                sheet = workbook.active # Get the default first sheet
                sheet.title = "GameLog"
                sheet.append(headers)
                for cell_h in sheet[1]: # First row (headers)
                    cell_h.font = Font(bold=True)
            
            # Append the new game data
            row_to_append = [game_data[h] for h in headers]
            sheet.append(row_to_append)
            print(f"Appended new game data. Sheet now has {sheet.max_row} data rows (excluding header if it was new).")

            # Auto-adjust column widths
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    if cell.value:
                        try:
                            cell_value_str = str(cell.value)
                            if len(cell_value_str) > max_length:
                                max_length = len(cell_value_str)
                        except:
                            pass
                
                adjusted_width = (max_length + 2) # Add a little padding
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(filename)
            print(f"Game data saved to {filename} with status: {self.game_end_condition}. Columns auto-adjusted.")

        except Exception as e:
            print(f"Error saving to Excel: {e}")
            messagebox.showerror("Excel Save Error", f"Could not save game data: {e}")

    def return_to_home(self, from_game_end_decline=False):
        if not from_game_end_decline and self.game_start_time is not None and not self.game_over:
            # Only save as "Quit" if game was active and not just finished & declined continuation
            self.game_end_condition = "Quit"
            self.save_game_to_excel()
        
        self.clear_screen()
        self.show_home_screen() # This also resets game_start_time, winner, game_over

    def quit_app(self):
        if self.game_start_time is not None and not self.game_over:
            self.game_end_condition = "Quit_From_App_Close" # Or just "Quit"
            self.save_game_to_excel()
        self.root.destroy()

    def restart_game(self):
        if self.game_start_time is not None and not self.game_over and self.winner == '':
            self.game_end_condition = "Reset"
            self.save_game_to_excel()
        
        self.setup_new_game()
        self.update_ui()

    def handle_click(self, event):
        if self.game_over:
            return

        click_col = event.x // self.SQUARE_SIZE
        click_row = event.y // self.SQUARE_SIZE
        click_coords = (click_col, click_row)

        # Ensure click is within board boundaries
        if not (0 <= click_col < self.BOARD_SIZE and 0 <= click_row < self.BOARD_SIZE):
            return

        current_player_pieces = []
        current_player_locations = []
        opponent_pieces = []
        opponent_locations = []
        player_color_string = ""

        if self.turn_step < 2: # White's turn (steps 0 or 1)
            current_player_pieces = self.white_pieces
            current_player_locations = self.white_locations
            opponent_pieces = self.black_pieces
            opponent_locations = self.black_locations
            player_color_string = 'white'
        else: # Black's turn (steps 2 or 3)
            current_player_pieces = self.black_pieces
            current_player_locations = self.black_locations
            opponent_pieces = self.white_pieces
            opponent_locations = self.white_locations
            player_color_string = 'black'
            
        if self.turn_step % 2 == 0:
            if click_coords in current_player_locations:
                self.selection = current_player_locations.index(click_coords)
                self.valid_moves = self.check_valid_moves()
                self.turn_step += 1
            else:
                self.selection = 100
                self.valid_moves = []
        
        else: # turn_step is 1 or 3
            if self.selection != 100: 
                if click_coords in self.valid_moves:
                    # Move the piece
                    moved_piece_index = self.selection
                    current_player_locations[moved_piece_index] = click_coords
                    self.total_moves_count += 1

                    # Check for capture
                    if click_coords in opponent_locations:
                        captured_idx = opponent_locations.index(click_coords)
                        captured_piece_type = opponent_pieces.pop(captured_idx)
                        opponent_locations.pop(captured_idx)
                        if player_color_string == 'white':
                            self.captured_pieces_white.append(captured_piece_type)
                        else:
                            self.captured_pieces_black.append(captured_piece_type)
                        
                        if captured_piece_type == 'king':
                            self.winner = player_color_string
                            self.game_over = True
                            self.game_end_condition = "King_Capture"
                            self.save_game_to_excel()
                            self.update_ui()
                            return
                    else:
                        if current_player_pieces[moved_piece_index] == 'pawn':
                            if (player_color_string == 'white' and click_coords[1] == 7) or \
                                (player_color_string == 'black' and click_coords[1] == 0):
                                current_player_pieces[moved_piece_index] = 'queen'

                    # Update all piece options for both players
                    self.white_options = self.check_options(self.white_pieces, self.white_locations, 'white')
                    self.black_options = self.check_options(self.black_pieces, self.black_locations, 'black')
                    
                    # Switch turn
                    self.turn_step = (self.turn_step + 1) % 4 # Cycles 0->1->2->3->0
                    self.selection = 100 # Deselect piece
                    self.valid_moves = []

                elif click_coords in current_player_locations: # Clicked another of own pieces
                    self.selection = current_player_locations.index(click_coords) # Select new piece
                    self.valid_moves = self.check_valid_moves()
                    # Keep turn_step as 1 or 3 (still this player's turn to make a move)
                else: # Clicked an invalid square (empty or unmovable)
                    self.selection = 100 # Deselect
                    self.valid_moves = []
                    self.turn_step -=1 # Go back to piece selection step (0 for White, 2 for Black)
            else: # No piece was selected, but turn_step was 1 or 3 (should not happen with proper logic)
                self.turn_step -=1 # Revert to selection state

        if self.game_over: # If king was captured
            self.update_ui() # Show the final move and capture
            self.process_game_over_prompt() # Then ask to continue
        else:
            self.update_ui()


if __name__ == "__main__":
    root = tk.Tk()
    game = ChessGame(root)
    root.mainloop()